import os
from google import genai
from google.genai import types
from backend.src.db import vortex as P


def _gather_client_documents(directory: str) -> str:
    """Walk memory/{company}/transcripts/ and concatenate readable docs."""
    context_text = ""
    for root, dirs, files in os.walk(directory):
        if any(skip in root.split(os.sep) for skip in ["abm_profiles"]):
            continue
        for file in files:
            if file.lower().endswith((".txt", ".md", ".csv", ".json")):
                filepath = os.path.join(root, file)
                try:
                    with open(filepath, "r", encoding="utf-8") as f:
                        context_text += f"\n--- DOCUMENT: {file} ---\n{f.read()}\n"
                except Exception:
                    pass
    return context_text


def _load_target_urls(targets_directory: str) -> list[str]:
    """Read urls.txt from targets/ and return a list of non-empty URLs."""
    urls_path = os.path.join(targets_directory, "urls.txt")
    if not os.path.isfile(urls_path):
        return []
    try:
        with open(urls_path, "r", encoding="utf-8") as f:
            return [line.strip() for line in f if line.strip() and not line.startswith("#")]
    except Exception:
        return []


def _parse_xlsx_files(targets_directory: str) -> str:
    """Read all .xlsx files in targets/ and convert to a text representation."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[MYDEI WARNING] openpyxl not installed — skipping .xlsx files.")
        return ""

    text_parts: list[str] = []
    if not os.path.isdir(targets_directory):
        return ""

    for file in sorted(os.listdir(targets_directory)):
        if not file.lower().endswith(".xlsx"):
            continue
        filepath = os.path.join(targets_directory, file)
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header = [str(c) if c is not None else "" for c in rows[0]]
                lines = [f"| {' | '.join(header)} |"]
                lines.append(f"| {' | '.join(['---'] * len(header))} |")
                for row in rows[1:]:
                    cells = [str(c) if c is not None else "" for c in row]
                    if all(c == "" for c in cells):
                        continue
                    lines.append(f"| {' | '.join(cells)} |")
                text_parts.append(
                    f"\n--- SPREADSHEET: {file} / Sheet: {sheet_name} ---\n"
                    + "\n".join(lines)
                )
            wb.close()
        except Exception as e:
            print(f"[MYDEI WARNING] Could not parse {file}: {e}")

    return "\n".join(text_parts)


class Mydei:
    """
    Mydei: The Autonomous ABM Target Researcher.

    Loads transcript context from memory/{company}/transcripts/ to understand
    the client's business, ICP, and what an ABM target would look like.
    Optionally parses .xlsx files from memory/{company}/targets/ to source
    a concrete list of potential targets.
    Uses Gemini with Google Search to prioritise and research ABM targets.
    """

    def __init__(self):
        self.api_key = os.environ.get("GEMINI_API_KEY")
        if self.api_key:
            self.client = genai.Client()
        else:
            self.client = None
            print("[MYDEI WARNING]: GEMINI_API_KEY environment variable is not set.")

    def generate_abm_briefing(self, company_keyword):
        if not self.client:
            return "[MYDEI WARNING] Gemini API key missing."

        transcripts_path = str(P.transcripts_dir(company_keyword))
        targets_path = str(P.targets_dir(company_keyword))

        if not os.path.exists(transcripts_path):
            return "[MYDEI WARNING] Client transcripts directory missing."

        # --- 1. Load transcript context (understanding the client) ---
        latest_path = os.path.join(transcripts_path, "latest.txt")
        used_fallback = False
        transcript_text = ""

        if os.path.isfile(latest_path):
            try:
                with open(latest_path, "r", encoding="utf-8") as f:
                    transcript_text = f.read()
            except Exception as e:
                return f"[MYDEI WARNING] Could not read latest.txt: {e}"

        if not transcript_text.strip():
            if os.path.isfile(latest_path):
                print(f"[MYDEI] latest.txt is empty for '{company_keyword}'; falling back to scanning client documents...")
            else:
                print(f"[MYDEI] latest.txt not found for '{company_keyword}'; falling back to scanning client documents...")
            transcript_text = _gather_client_documents(transcripts_path)
            used_fallback = True

        if not transcript_text.strip():
            return (
                f"[MYDEI WARNING] No readable context (latest.txt missing/empty "
                f"and no documents under memory/{company_keyword}/transcripts/)."
            )

        # --- 2. Load spreadsheet data and reference URLs from targets/ ---
        spreadsheet_text = _parse_xlsx_files(targets_path)
        has_spreadsheets = bool(spreadsheet_text.strip())

        target_urls = _load_target_urls(targets_path)

        # --- 3. Configure Gemini with Google Search ---
        config = types.GenerateContentConfig(
            temperature=0.3,
            tools=[{"google_search": {}}],
        )

        # --- 4. Build the prompt ---
        if used_fallback:
            transcript_label = "CLIENT DOCUMENTS (full-folder scan — latest.txt was missing or empty)"
        else:
            transcript_label = "LATEST INTERVIEW (latest.txt — the client's most recent interview)"

        transcript_block = f"{transcript_label}:\n{transcript_text}"

        url_block = ""
        if target_urls:
            url_list = "\n".join(f"  - {u}" for u in target_urls)
            url_block = f"""
REFERENCE URLS (from memory/{company_keyword}/targets/urls.txt):
The following URLs have been provided as relevant industry articles, news, or resources.
Use your Google Search tool to access and read each URL. Extract any information that is
useful for understanding the client's market, identifying ABM targets, or finding tactical
angles for ABM posts. Incorporate relevant findings into your target profiles.

{url_list}
"""

        if has_spreadsheets:
            spreadsheet_block = (
                f"SPREADSHEET DATA (from memory/{company_keyword}/targets/):\n"
                f"{spreadsheet_text}"
            )
            prompt = f"""
You are Mydei, an expert Account-Based Marketing (ABM) researcher.

STEP 1 — UNDERSTAND THE CLIENT:
Read the following transcripts and documents carefully. Understand what the client does, who they serve,
what their ideal customer profile (ICP) looks like, and what pain points their product or service solves.
Build a clear mental model of the client's ICP before proceeding.

{transcript_block}

STEP 2 — FILTER AND SHORTLIST:
Below is spreadsheet data containing potential ABM targets. Scan the ENTIRE dataset — it may contain
hundreds of entries. Using your understanding of the client's ICP from Step 1, rigorously filter
this list down to a shortlist of the strongest candidates. Apply these filters:

  a) ICP FIT — Does this target's industry, size, role, or strategic direction match what the
     client actually sells to? Reject targets that fall outside the client's ICP.
  b) RELEVANCE — Does the target have pain points that the client's product or service directly
     addresses? Generic or tangential matches are not good enough.
  c) ACCESSIBILITY — Is the target a real, active company or executive? Defunct companies,
     placeholder entries, or unverifiable names should be discarded.

Output your shortlist (up to 10 names) with a one-line rationale for each before proceeding.

{spreadsheet_block}
{url_block}
STEP 3 — VERIFY AND PROFILE:
From your shortlist, pick the top 6 targets. For EACH one, use your Google Search tool to:
  1. Verify the target actually exists and is currently active.
  2. Confirm they plausibly fit the client's ICP based on real-world evidence.
  3. Research their recent activity for tactical ABM angles.

DISCARD any target that fails verification (does not exist, is defunct, or turns out not to fit
the ICP upon closer inspection). Only profile verified, ICP-fit targets.

If after filtering and verification NO viable targets remain, fall back to identifying targets
from the transcripts themselves. If there are still none, reply EXACTLY with: "NO ABM TARGETS FOUND."

For each verified target, provide a Markdown profile:
### [Target Name]
1. Target Overview — who they are and why they are a high-priority target for this client
2. Recent News & Strategic Shifts (last 6-12 months)
3. Predicted Pain Points — specifically as they relate to what this client can solve
4. Ingress Strategy (2 unique angles to catch their attention in a LinkedIn post)
"""
        else:
            prompt = f"""
You are Mydei, an expert Account-Based Marketing (ABM) researcher.

STEP 1 — UNDERSTAND THE CLIENT:
Read the following transcripts and documents. Understand what the client does, who they serve,
what their ideal customer profile (ICP) looks like, and what pain points their product or service solves.

{transcript_block}
{url_block}
STEP 2 — IDENTIFY TARGETS:
From the transcripts (and reference URLs if provided), identify any specific companies, brands, or executives explicitly mentioned
or heavily implied as ideal prospects, clients, or ABM targets. Only select targets that genuinely
fit the client's ICP — not just any company name that appears in conversation.

If you CANNOT find any specific named companies or individuals that fit the client's ICP,
reply EXACTLY with: "NO ABM TARGETS FOUND."

STEP 3 — VERIFY AND PROFILE:
If you found targets, pick up to 4 of the most promising ones. For each, use your Google Search
tool to verify they actually exist and are currently active, confirm they fit the client's ICP,
and research them for tactical ABM angles. Discard any target that fails verification.

For each verified target, provide a Markdown profile:
### [Target Name]
1. Target Overview — who they are and why they are a high-priority target for this client
2. Recent News & Strategic Shifts (last 6-12 months)
3. Predicted Pain Points — specifically as they relate to what this client can solve
4. Ingress Strategy (2 unique angles to catch their attention in a LinkedIn post)
"""

        try:
            src_label = "client documents (fallback)" if used_fallback else "latest.txt"
            xlsx_label = f" + {spreadsheet_text.count('SPREADSHEET:')} spreadsheet(s)" if has_spreadsheets else ""
            urls_label = f" + {len(target_urls)} URL(s)" if target_urls else ""
            print(f"[MYDEI] Using {src_label}{xlsx_label}{urls_label} for '{company_keyword}' to extract and research ABM targets...")
            response = self.client.models.generate_content(
                model='gemini-3.1-pro-preview',
                contents=prompt,
                config=config,
            )
            return response.text
        except Exception as e:
            print(f"Mydei API Error: {e}")
            return "[MYDEI WARNING] API Error."
